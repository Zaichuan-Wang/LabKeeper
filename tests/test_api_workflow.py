"""FastAPI 集成流程测试，使用临时 SQLite 数据库。"""

from datetime import datetime, timedelta


def api_ok(response, status_code=200):
    assert response.status_code == status_code, response.text
    return response.json()


def antibody_fields(target="CD45", conjugate="APC"):
    return {"antibody_target": target, "antibody_conjugate": conjugate}


def test_save_new_brand_option_is_explicit(app_client, auth_headers):
    from services import options_config

    options_config.save_dropdown_options({"brands": ["CST"]})

    api_ok(
        app_client.post(
            "/api/orders",
            headers=auth_headers,
            json={"name": "No save brand", "category": "试剂盒", "brand": "NewCo", "quantity": 1},
        ),
        201,
    )
    assert "NewCo" not in options_config.load_dropdown_options()["brands"]

    saved_order = api_ok(
        app_client.post(
            "/api/orders",
            headers=auth_headers,
            json={
                "name": "Save brand",
                "category": "试剂盒",
                "brand": "NewCo",
                "save_brand_option": True,
                "quantity": 1,
            },
        ),
        201,
    )
    assert "NewCo" in saved_order["options"]["brands"]
    assert options_config.load_dropdown_options()["brands"].count("NewCo") == 1

    reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={"item_type": "reagent", "name": "Edit brand", "category": "其他", "brand": "CST", "quantity": 1},
        ),
        201,
    )["item"]
    updated = api_ok(
        app_client.patch(
            f"/api/inventory/item?item_type=reagent&id={reagent['id']}",
            headers=auth_headers,
            json={"brand": "EditCo", "save_brand_option": True},
        )
    )
    assert "EditCo" in updated["options"]["brands"]

    api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Duplicate brand",
                "category": "其他",
                "brand": "newco",
                "save_brand_option": True,
                "quantity": 1,
            },
        ),
        201,
    )
    assert options_config.load_dropdown_options()["brands"].count("NewCo") == 1


def test_regular_user_can_update_dropdown_settings(app_client, auth_headers):
    from services import options_config

    created = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={
                "username": "options-user",
                "password": "secret123",
                "display_name": "Options User",
                "role": "user",
                "permissions": {},
            },
        ),
        201,
    )["item"]
    assert created["role"] == "user"

    login = api_ok(app_client.post("/api/login", json={"username": "options-user", "password": "secret123"}))
    user_headers = {"Authorization": f"Bearer {login['token']}"}
    updated = api_ok(
        app_client.patch(
            "/api/settings/dropdowns",
            headers=user_headers,
            json={"brands": ["CST", "OpenBrand"]},
        )
    )["item"]

    assert "OpenBrand" in updated["brands"]
    assert "OpenBrand" in options_config.load_dropdown_options()["brands"]


def test_catalog_usage_warns_when_catalog_already_exists(app_client, auth_headers):
    first = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={"item_type": "reagent", "name": "Catalog One", "category": "其他", "catalog_no": "AI-CAT-1", "quantity": 1},
        ),
        201,
    )["item"]
    api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={"item_type": "reagent", "name": "Catalog Two", "category": "其他", "catalog_no": "AI-CAT-1", "quantity": 1},
        ),
        201,
    )

    usage = api_ok(app_client.get("/api/inventory/catalog-usage?catalog_no=AI-CAT-1", headers=auth_headers))
    assert usage["has_existing"] is True
    assert usage["count"] == 2
    assert {item["name"] for item in usage["items"]} == {"Catalog One", "Catalog Two"}

    excluded = api_ok(app_client.get(f"/api/inventory/catalog-usage?catalog_no=AI-CAT-1&exclude_id={first['id']}", headers=auth_headers))
    assert excluded["has_existing"] is True
    assert excluded["count"] == 1
    assert excluded["items"][0]["name"] == "Catalog Two"

    empty = api_ok(app_client.get("/api/inventory/catalog-usage?catalog_no=UNKNOWN-CAT", headers=auth_headers))
    assert empty["has_existing"] is False
    assert empty["count"] == 0


def test_inventory_storage_and_permission_workflow(app_client, auth_headers):
    health = api_ok(app_client.get("/api/health"))
    assert health["ok"] is True
    assert app_client.get("/api/dashboard").status_code == 200

    dashboard = api_ok(app_client.get("/api/dashboard", headers=auth_headers))
    assert dashboard["metrics"]["total_inventory"] == 0

    tree = api_ok(app_client.get("/api/storage/tree", headers=auth_headers))
    roots = [item for item in tree["items"] if item.get("parent_id") is None]
    assert len(roots) == 1
    root_id = roots[0]["id"]

    freezer = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": root_id, "name": "Freezer-1", "space_type": "2", "rows": 2, "cols": 3},
        ),
        201,
    )["item"]
    assert freezer["space_type"] == 2
    freezer_update = api_ok(
        app_client.patch(
            f"/api/storage/nodes/{freezer['id']}",
            headers=auth_headers,
            json={"space_type": "1"},
        )
    )["item"]
    assert freezer_update["space_type"] == 1
    rack = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": freezer["id"], "name": "Rack-A", "rows": 4, "cols": 4},
        ),
        201,
    )["item"]
    frame_space = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": rack["id"], "name": "Frame-A", "rows": 9, "cols": 9},
        ),
        201,
    )["item"]

    forgiving_space = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": freezer["id"], "name": "Forgiving-space", "sort_order": ""},
        ),
        201,
    )["item"]
    assert forgiving_space["sort_order"] == 0

    sample_result = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "sample",
                "name": "P-SMOKE-001",
                "category": "组织",
                "tube_count": 2,
                "separate_items": True,
                "amount": 0.5,
                "amount_unit": "mL",
                "status": "可用",
                "storage_node_id": frame_space["id"],
                "grid_cell": "A1",
            },
        ),
        201,
    )
    assert sample_result["count"] == 2
    sample_items = sample_result["items"]
    assert [item["grid_cell"] for item in sample_items] == ["A1", "A2"]

    reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Ab-test",
                "category": "抗体",
                "brand": "CST",
                "catalog_no": "SMOKE-CAT",
                **antibody_fields("CD45", "APC"),
                "quantity": 1,
                "status": "可用",
                "storage_node_id": frame_space["id"],
                "grid_cell": "B1",
            },
        ),
        201,
    )["item"]

    frame_visual = api_ok(app_client.get(f"/api/storage/visual?node_id={frame_space['id']}", headers=auth_headers))
    occupied = {item["grid_cell"]: item["code"] for item in frame_visual["frame_items"]}
    assert sample_items[0]["code"] in occupied["A1"]
    assert reagent["code"] in occupied["B1"]

    reagent_only_user = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={
                "username": "reagent-only",
                "password": "secret123",
                "display_name": "Reagent Only",
                "role": "user",
                "permissions": {
                    "inventory.search": True,
                    "inventory.view_reagents": True,
                    "inventory.view_samples": False,
                },
            },
        ),
        201,
    )["item"]
    assert reagent_only_user["permissions"]["inventory.view_reagents"] is True
    assert reagent_only_user["permissions"]["inventory.view_samples"] is False
    reagent_only_login = api_ok(app_client.post("/api/login", json={"username": "reagent-only", "password": "secret123"}))
    reagent_only_headers = {"Authorization": f"Bearer {reagent_only_login['token']}"}

    reagent_only_visual = api_ok(app_client.get(f"/api/storage/visual?node_id={frame_space['id']}", headers=reagent_only_headers))
    assert all(item["item_type"] == "reagent" for item in reagent_only_visual["frame_items"])
    assert reagent["id"] in {item["id"] for item in reagent_only_visual["frame_items"]}
    assert not any(item["item_type"] == "sample" and item["id"] == sample_items[0]["id"] for item in reagent_only_visual["direct_items"])
    assert reagent_only_visual["stats"]["direct"] == 1

    reagent_only_search = api_ok(
        app_client.get("/api/inventory/search?item_type=all&keyword=P-SMOKE-001&purpose=global", headers=reagent_only_headers)
    )
    assert all(item["item_type"] != "sample" for item in reagent_only_search["items"])
    denied_sample_search = api_ok(
        app_client.get("/api/inventory/search?item_type=sample&keyword=P-SMOKE-001&purpose=global", headers=reagent_only_headers)
    )
    assert denied_sample_search["total"] == 0
    assert denied_sample_search["items"] == []
    assert app_client.get(f"/api/inventory/item?item_type=sample&id={sample_items[0]['id']}", headers=reagent_only_headers).status_code == 403
    assert app_client.post(
        "/api/checkouts",
        headers=reagent_only_headers,
        json={"item_type": "sample", "item_id": sample_items[0]["id"], "reason": "no sample permission"},
    ).status_code == 403
    assert app_client.post(
        "/api/aliquots",
        headers=reagent_only_headers,
        json={"item_type": "sample", "source_item_id": sample_items[0]["id"], "tube_count": 1},
    ).status_code == 403

    maintainer = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={
                "username": "maintainer",
                "password": "secret123",
                "display_name": "Maintainer",
                "role": "user",
                "permissions": {"inventory.manage": True, "inventory.view_reagents": True, "inventory.view_samples": True},
            },
        ),
        201,
    )["item"]
    assert maintainer["permissions"]["inventory.manage"] is True
    maintainer_login = api_ok(app_client.post("/api/login", json={"username": "maintainer", "password": "secret123"}))
    maintainer_headers = {"Authorization": f"Bearer {maintainer_login['token']}"}
    maintainer_aliquot_search = api_ok(
        app_client.get(
            "/api/inventory/search?item_type=sample&keyword=P-SMOKE-001&available=1&purpose=aliquot",
            headers=maintainer_headers,
        )
    )
    assert any(item["id"] == sample_items[0]["id"] for item in maintainer_aliquot_search["items"])
    maintainer_aliquot = api_ok(
        app_client.post(
            "/api/aliquots",
            headers=maintainer_headers,
            json={"item_type": "sample", "source_item_id": sample_items[0]["id"], "tube_count": 1},
        ),
        201,
    )
    assert maintainer_aliquot["count"] == 1

    shrink_sample = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "sample",
                "name": "P-SHRINK-001",
                "category": "组织",
                "tube_count": 1,
                "status": "可用",
                "storage_node_id": frame_space["id"],
                "grid_cell": "A9",
            },
        ),
        201,
    )["item"]
    shrink_result = api_ok(
        app_client.patch(
            f"/api/storage/nodes/{frame_space['id']}",
            headers=auth_headers,
            json={"cols": 8},
        )
    )
    assert shrink_result["cleared_out_of_bounds"]["samples"] == 1
    shrink_detail = api_ok(
        app_client.get(f"/api/inventory/item?item_type=sample&id={shrink_sample['id']}", headers=auth_headers)
    )["item"]
    assert shrink_detail["storage_node_id"] == frame_space["id"]
    assert shrink_detail["grid_cell"] in ("", None)
    frame_visual = api_ok(app_client.get(f"/api/storage/visual?node_id={frame_space['id']}", headers=auth_headers))
    direct_ids_without_position = {
        item["id"] for item in frame_visual["direct_items"]
        if item["item_type"] == "sample" and not item.get("grid_cell")
    }
    assert shrink_sample["id"] in direct_ids_without_position

    freezer_search = api_ok(
        app_client.get(f"/api/inventory/search?item_type=reagent&storage_node_id={freezer['id']}", headers=auth_headers)
    )
    assert any(item["id"] == reagent["id"] for item in freezer_search["items"])

    path_keyword_search = api_ok(
        app_client.get("/api/inventory/search?item_type=all&keyword=Freezer-1", headers=auth_headers)
    )
    assert any(item["item_type"] == "reagent" and item["id"] == reagent["id"] for item in path_keyword_search["items"])

    keyword_search = api_ok(
        app_client.get("/api/inventory/search?item_type=reagent&keyword=SMOKE-CAT&page=1&page_size=1", headers=auth_headers)
    )
    assert keyword_search["total"] >= 1
    assert keyword_search["page"] == 1
    assert keyword_search["page_size"] == 1
    assert keyword_search["items"][0]["id"] == reagent["id"]

    bad_timeline = app_client.get("/api/inventory/timeline?item_type=reagent&id=bad", headers=auth_headers)
    assert bad_timeline.status_code == 400

    sample_search = api_ok(
        app_client.get("/api/inventory/search?item_type=sample&keyword=P-SMOKE-001&available=1", headers=auth_headers)
    )
    assert sample_search["count"] >= 2

    unplaced_reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Unplaced reagent",
                "category": "抗体",
                **antibody_fields("CD3", "FITC"),
                "quantity": 1,
                "status": "可用",
            },
        ),
        201,
    )["item"]
    assert unplaced_reagent["storage_node_id"] == -3
    assert unplaced_reagent["storage_location"] == "未归位"

    unplaced_sample = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "sample",
                "name": "P-UNPLACED-001",
                "category": "血清",
                "tube_count": 1,
                "status": "可用",
            },
        ),
        201,
    )["item"]
    assert unplaced_sample["storage_node_id"] == -3
    assert unplaced_sample["storage_location"] == "未归位"

    unplaced_visual = api_ok(app_client.get("/api/storage/visual?node_id=-3", headers=auth_headers))
    assert unplaced_visual["current"]["id"] == -3
    assert unplaced_visual["current"]["name"] == "未归位"
    assert {item["item_type"] for item in unplaced_visual["direct_items"]} >= {"reagent", "sample"}

    moved = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={"item_type": "sample", "item_id": sample_items[1]["id"], "to_storage_node_id": freezer["id"], "reason": "整理库存"},
        ),
        201,
    )["item"]
    assert moved["to_storage_node_id"] == freezer["id"]

    import db.database as database

    with database.connect() as conn:
        movement_count = conn.execute("SELECT COUNT(*) AS n FROM movements").fetchone()["n"]
    unchanged_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={"item_type": "sample", "item_id": sample_items[1]["id"], "to_storage_node_id": freezer["id"], "reason": "拖拽移动"},
        ),
        201,
    )["item"]
    assert unchanged_move["unchanged"] is True
    assert unchanged_move["to_storage_node_id"] == freezer["id"]
    assert unchanged_move["can_rollback"] is False
    with database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM movements").fetchone()["n"] == movement_count

    validation = api_ok(
        app_client.post(
            "/api/validations",
            headers=auth_headers,
            json={
                "catalog_no": "SMOKE-CAT",
                "validation_date": "2026-06-01",
                "method": "流式",
                "result": "通过",
                "description": "阳性对照清楚，背景低。",
                "image_path": "data/validation_images/example.png",
            },
        ),
        201,
    )["item"]
    assert validation["result"] == "通过"
    api_ok(
        app_client.post(
            "/api/validations",
            headers=auth_headers,
            json={"catalog_no": "SMOKE-CAT", "validation_date": "2026-06-02", "method": "流式", "result": "不通过"},
        ),
        201,
    )
    api_ok(
        app_client.post(
            "/api/validations",
            headers=auth_headers,
            json={"catalog_no": "SMOKE-CAT", "validation_date": "2026-06-03", "method": "流式", "result": "待复核"},
        ),
        201,
    )
    reagent_after_validations = api_ok(
        app_client.get(f"/api/inventory/item?item_type=reagent&id={reagent['id']}", headers=auth_headers)
    )["item"]
    assert reagent_after_validations["validation_status"] == "通过"
    validation_filter = api_ok(
        app_client.get("/api/inventory/search?item_type=reagent&validation_status=通过", headers=auth_headers)
    )
    assert any(item["id"] == reagent["id"] for item in validation_filter["items"])

    reagent_timeline = api_ok(
        app_client.get(f"/api/inventory/timeline?item_type=reagent&id={reagent['id']}", headers=auth_headers)
    )
    assert all(event["event_type"] != "validation" for event in reagent_timeline["items"])
    reagent_detail = api_ok(
        app_client.get(f"/api/inventory/item?item_type=reagent&id={reagent['id']}", headers=auth_headers)
    )
    assert any(row["catalog_no"] == "SMOKE-CAT" for row in reagent_detail["validations"])

    checkout = api_ok(
        app_client.post(
            "/api/checkouts",
            headers=auth_headers,
            json={"item_type": "sample", "item_id": sample_items[0]["id"], "reason": "实验消耗"},
        ),
        201,
    )["item"]
    assert checkout["to_location"] == "已出库"

    limited_user = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={
                "username": "limited",
                "password": "secret123",
                "display_name": "Limited User",
                "role": "user",
                "permissions": {"inventory.manage": False, "location.manage": False, "inventory.search": False},
            },
        ),
        201,
    )["item"]
    assert limited_user["permissions"]["inventory.search"] is False

    limited_login = api_ok(app_client.post("/api/login", json={"username": "limited", "password": "secret123"}))
    limited_headers = {"Authorization": f"Bearer {limited_login['token']}"}

    denied_search = app_client.get("/api/inventory/search?item_type=all&keyword=Ab-test&purpose=global", headers=limited_headers)
    assert denied_search.status_code == 403
    for path in (
        "/api/orders",
        "/api/validations",
        "/api/checkouts",
        "/api/movements",
    ):
        denied_history = app_client.get(path, headers=limited_headers)
        assert denied_history.status_code == 403
    denied_move = app_client.post(
        "/api/movements",
        headers=limited_headers,
        json={"item_type": "reagent", "item_id": reagent["id"], "to_storage_node_id": freezer["id"]},
    )
    assert denied_move.status_code == 403
    limited_reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=limited_headers,
            json={
                "item_type": "reagent",
                "name": "Limited reagent",
                "category": "抗体",
                **antibody_fields("CD19", "PE"),
                "quantity": 1,
                "status": "可用",
            },
        ),
        201,
    )["item"]
    denied_reagent_edit = app_client.patch(
        f"/api/inventory/item?item_type=reagent&id={limited_reagent['id']}",
        headers=limited_headers,
        json={"name": "Limited reagent edit"},
    )
    assert denied_reagent_edit.status_code == 403
    limited_sample = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=limited_headers,
            json={"item_type": "sample", "name": "LIMIT-SAMPLE-001", "category": "血清", "tube_count": 1, "status": "可用"},
        ),
        201,
    )["item"]
    denied_limited_aliquot_search = app_client.get(
        "/api/inventory/search?item_type=sample&keyword=LIMIT-SAMPLE-001&available=1&purpose=aliquot",
        headers=limited_headers,
    )
    assert denied_limited_aliquot_search.status_code == 403
    denied_limited_aliquot = app_client.post(
        "/api/aliquots",
        headers=limited_headers,
        json={"item_type": "sample", "source_item_id": limited_sample["id"], "tube_count": 1},
    )
    assert denied_limited_aliquot.status_code == 403

    occupied_delete = app_client.delete(f"/api/storage/nodes/{frame_space['id']}", headers=auth_headers)
    assert occupied_delete.status_code == 400

    limited_order = api_ok(
        app_client.post(
            "/api/orders",
            headers=limited_headers,
            json={"name": "Limited-order", "category": "其他", "catalog_no": "LIMIT-CAT", "quantity": 1, "price": 88.5},
        ),
        201,
    )["item"]
    assert limited_order["name"] == "Limited-order"
    limited_order_candidates = api_ok(app_client.get("/api/orders?purpose=form", headers=limited_headers))["items"]
    assert any(item["id"] == limited_order["id"] for item in limited_order_candidates)

    arrival = api_ok(
        app_client.post(
            "/api/arrivals",
            headers=auth_headers,
            json={"order_id": limited_order["id"], "arrival_quantity": 1},
        ),
        201,
    )
    assert arrival["items"][0]["storage_node_id"] == -3
    assert arrival["items"][0]["storage_location"] == "未归位"
    order_ledger = api_ok(app_client.get("/api/orders", headers=auth_headers))
    ledger_row = next(item for item in order_ledger["items"] if item["id"] == limited_order["id"])
    assert ledger_row["arrival_status"] == "已到货"
    assert ledger_row["arrival_count"] == 1
    assert ledger_row["arrived_quantity"] == 1
    assert ledger_row["arrival_codes"] == f"RG{arrival['item_id']:06d}"
    assert ledger_row["arrival_locations"] == "未归位"
    linked_detail = api_ok(
        app_client.get(f"/api/inventory/item?item_type=reagent&id={arrival['item_id']}", headers=auth_headers)
    )["item"]
    assert linked_detail["price"] == 88.5
    assert "linked_order_price" not in linked_detail

    empty_bin = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": freezer["id"], "name": "Empty-bin"},
        ),
        201,
    )["item"]

    with database.connect() as conn:
        conn.execute(
            """
            INSERT INTO movements
                (object_type, object_id, item_type, item_id, from_storage_node_id, from_grid_cell,
                 to_storage_node_id, to_grid_cell, from_location_snapshot, to_location_snapshot, moved_by, moved_at, reason, note)
            VALUES ('试剂', ?, 'reagent', ?, ?, 'A1', ?, 'B1', 'Empty-bin / A1', 'Empty-bin / B1', 1, '2026-06-01 10:00:00', '历史引用测试', '')
            """,
            (reagent["code"], reagent["id"], empty_bin["id"], empty_bin["id"]),
        )
        conn.commit()

    deleted = api_ok(app_client.delete(f"/api/storage/nodes/{empty_bin['id']}", headers=auth_headers))
    assert deleted["cleared_history_refs"] == {"movement_refs": 2}
    with database.connect() as conn:
        movement = conn.execute("SELECT from_storage_node_id, to_storage_node_id FROM movements ORDER BY id DESC LIMIT 1").fetchone()
        assert movement["from_storage_node_id"] == -3
        assert movement["to_storage_node_id"] == -3


def test_admin_user_initial_password_and_reset(app_client, auth_headers, monkeypatch):
    from services import admin

    monkeypatch.setattr(admin.config, "INITIAL_PASSWORD", "lab-init-789")
    created = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={"username": "initial-user", "display_name": "Initial User", "role": "user", "permissions": {}},
        ),
        201,
    )["item"]

    login = api_ok(app_client.post("/api/login", json={"username": "initial-user", "password": "lab-init-789"}))
    assert login["user"]["username"] == "initial-user"

    api_ok(
        app_client.patch(
            f"/api/users/{created['id']}",
            headers=auth_headers,
            json={"password": "changed-789"},
        )
    )
    changed_login = api_ok(app_client.post("/api/login", json={"username": "initial-user", "password": "changed-789"}))
    assert changed_login["user"]["username"] == "initial-user"

    api_ok(app_client.post(f"/api/users/{created['id']}/reset-password", headers=auth_headers))
    reset_login = api_ok(app_client.post("/api/login", json={"username": "initial-user", "password": "lab-init-789"}))
    assert reset_login["user"]["username"] == "initial-user"


def test_movement_merge_window_merges_recent_moves_but_not_orders(app_client, auth_headers):
    import db.database as database

    tree = api_ok(app_client.get("/api/storage/tree", headers=auth_headers))
    root_id = next(item["id"] for item in tree["items"] if item.get("parent_id") is None)
    freezer = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": root_id, "name": "Merge-Freezer"},
        ),
        201,
    )["item"]
    rack = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": freezer["id"], "name": "Merge-Rack"},
        ),
        201,
    )["item"]
    slot_a = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": rack["id"], "name": "Slot-A"},
        ),
        201,
    )["item"]
    slot_b = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": rack["id"], "name": "Slot-B"},
        ),
        201,
    )["item"]

    reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Merge-Reagent",
                "category": "抗体",
                "catalog_no": "MERGE-CAT",
                **antibody_fields("CD4", "APC"),
                "quantity": 1,
                "status": "可用",
                "storage_node_id": slot_a["id"],
            },
        ),
        201,
    )["item"]

    first_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": reagent["id"],
                "to_storage_node_id": slot_b["id"],
                "reason": "库存移动",
                "note": "第一次移动",
            },
        ),
        201,
    )["item"]
    assert first_move.get("merged") is not True
    assert first_move["reason"] == "位置移动"
    assert first_move["to_storage_node_id"] == slot_b["id"]
    assert first_move["note"] == "第一次移动"

    with database.connect() as conn:
        movement_count = conn.execute("SELECT COUNT(*) AS n FROM movements WHERE item_type = 'reagent' AND item_id = ?", (reagent["id"],)).fetchone()["n"]
    assert movement_count == 2

    second_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": reagent["id"],
                "to_storage_node_id": slot_a["id"],
                "reason": "库存移动",
                "note": "第二次移动",
            },
        ),
        201,
    )["item"]
    assert second_move["merged"] is True
    assert second_move["unchanged"] is True
    assert second_move["deleted"] is True
    with database.connect() as conn:
        movement_rows = conn.execute(
            "SELECT reason, from_storage_node_id, to_storage_node_id FROM movements WHERE item_type = 'reagent' AND item_id = ? ORDER BY id",
            (reagent["id"],),
        ).fetchall()
    assert len(movement_rows) == 1
    assert movement_rows[0]["reason"] == "入库登记"

    ordered = api_ok(
        app_client.post(
            "/api/orders",
            headers=auth_headers,
            json={"name": "Merge-Order", "category": "其他", "catalog_no": "MERGE-ORDER", "quantity": 1, "price": 12.3},
        ),
        201,
    )["item"]
    arrival = api_ok(
        app_client.post(
            "/api/arrivals",
            headers=auth_headers,
            json={"order_id": ordered["id"], "storage_node_id": slot_a["id"]},
        ),
        201,
    )["item"]
    assert arrival["storage_node_id"] == slot_a["id"]

    arrival_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": arrival["item_id"],
                "to_storage_node_id": slot_b["id"],
                "reason": "整理位置",
            },
        ),
        201,
    )["item"]
    assert arrival_move["merged"] is True
    assert arrival_move["reason"] == "到货入库"
    assert arrival_move["to_storage_node_id"] == slot_b["id"]

    with database.connect() as conn:
        timeline_rows = conn.execute(
            "SELECT reason, to_storage_node_id, to_location_snapshot FROM movements WHERE item_type = 'reagent' AND item_id = ? ORDER BY id",
            (arrival["item_id"],),
        ).fetchall()
    assert [row["reason"] for row in timeline_rows] == ["订购", "到货入库"]
    assert timeline_rows[-1]["to_storage_node_id"] == slot_b["id"]

    order_rows = api_ok(app_client.get("/api/orders?catalog_no=MERGE-ORDER", headers=auth_headers))["items"]
    assert order_rows[0]["arrival_status"] == "已到货"

    api_ok(
        app_client.patch(
            "/api/settings/dropdowns",
            headers=auth_headers,
            json={"movement_merge_window_minutes": 0},
        )
    )
    reagent2 = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "No-Merge-Reagent",
                "category": "抗体",
                "catalog_no": "NO-MERGE",
                **antibody_fields("CD8", "FITC"),
                "quantity": 1,
                "status": "可用",
                "storage_node_id": slot_a["id"],
            },
        ),
        201,
    )["item"]
    no_merge = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": reagent2["id"],
                "to_storage_node_id": slot_b["id"],
                "reason": "库存移动",
            },
        ),
        201,
    )["item"]
    assert no_merge.get("merged") is not True
    with database.connect() as conn:
        assert conn.execute(
            "SELECT COUNT(*) AS n FROM movements WHERE item_type = 'reagent' AND item_id = ?",
            (reagent2["id"],),
        ).fetchone()["n"] == 2


def test_movement_delete_requires_latest_movement(app_client, auth_headers):
    import db.database as database

    tree = api_ok(app_client.get("/api/storage/tree", headers=auth_headers))
    root_id = next(item["id"] for item in tree["items"] if item.get("parent_id") is None)
    freezer = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": root_id, "name": "Delete-Move-Freezer"},
        ),
        201,
    )["item"]
    rack = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": freezer["id"], "name": "Delete-Move-Rack"},
        ),
        201,
    )["item"]
    slot_a = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": rack["id"], "name": "Delete-Move-A"},
        ),
        201,
    )["item"]
    slot_b = api_ok(
        app_client.post(
            "/api/storage/nodes",
            headers=auth_headers,
            json={"parent_id": rack["id"], "name": "Delete-Move-B"},
        ),
        201,
    )["item"]

    reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Delete-Move-Reagent",
                "category": "抗体",
                "catalog_no": "DEL-MOVE",
                **antibody_fields("CD20", "BV421"),
                "quantity": 1,
                "status": "可用",
                "storage_node_id": slot_a["id"],
            },
        ),
        201,
    )["item"]
    api_ok(
        app_client.patch(
            "/api/settings/dropdowns",
            headers=auth_headers,
            json={"movement_merge_window_minutes": 0},
        )
    )

    first_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": reagent["id"],
                "to_storage_node_id": slot_b["id"],
                "reason": "库存移动",
            },
        ),
        201,
    )["item"]
    second_move = api_ok(
        app_client.post(
            "/api/movements",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "item_id": reagent["id"],
                "to_storage_node_id": slot_a["id"],
                "reason": "库存移动",
            },
        ),
        201,
    )["item"]

    denied = app_client.post(f"/api/movements/{first_move['id']}/rollback", headers=auth_headers)
    assert denied.status_code == 409
    assert "之后还有移动记录" in denied.text
    timeline = api_ok(app_client.get(f"/api/inventory/timeline?item_type=reagent&id={reagent['id']}", headers=auth_headers))
    movement_events = {item["related_id"]: item for item in timeline["items"] if item["related_table"] == "movements"}
    assert movement_events[first_move["id"]]["details"]["can_rollback"] is False
    assert movement_events[second_move["id"]]["details"]["can_rollback"] is True

    deleted = api_ok(app_client.post(f"/api/movements/{second_move['id']}/rollback", headers=auth_headers), 201)
    assert deleted["deleted_id"] == second_move["id"]
    with database.connect() as conn:
        movement_count = conn.execute("SELECT COUNT(*) AS n FROM movements WHERE item_type = 'reagent' AND item_id = ?", (reagent["id"],)).fetchone()["n"]
        current = conn.execute("SELECT storage_node_id FROM reagents WHERE id = ?", (reagent["id"],)).fetchone()
    assert movement_count == 2
    assert current["storage_node_id"] == slot_a["id"]


def test_emergency_excel_export_skips_internal_tables(app_client, auth_headers):
    from io import BytesIO

    from openpyxl import load_workbook

    table_data = api_ok(app_client.get("/api/excel/tables", headers=auth_headers))
    table_names = {item["name"] for item in table_data["items"]}
    assert "reagents" in table_names
    assert "schema_migrations" not in table_names

    response = app_client.get("/api/excel/export?mode=data&limit=5", headers=auth_headers)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert "reagents" in workbook.sheetnames
    assert "schema_migrations" not in workbook.sheetnames


def test_validation_edit_permissions_owner_and_admin(app_client, auth_headers):
    owner = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={"username": "owner", "password": "owner123", "display_name": "Owner", "role": "user", "permissions": {}},
        ),
        201,
    )["item"]
    other = api_ok(
        app_client.post(
            "/api/users",
            headers=auth_headers,
            json={"username": "other", "password": "other123", "display_name": "Other", "role": "user", "permissions": {}},
        ),
        201,
    )["item"]
    assert owner["role"] == "user"
    assert other["role"] == "user"

    owner_login = api_ok(app_client.post("/api/login", json={"username": "owner", "password": "owner123"}))
    other_login = api_ok(app_client.post("/api/login", json={"username": "other", "password": "other123"}))
    owner_headers = {"Authorization": f"Bearer {owner_login['token']}"}
    other_headers = {"Authorization": f"Bearer {other_login['token']}"}

    created = api_ok(
        app_client.post(
            "/api/validations",
            headers=owner_headers,
            json={"catalog_no": "VAL-EDIT-CAT", "validation_date": "2026-06-16", "method": "WB", "result": "待复核", "description": "初稿"},
        ),
        201,
    )["item"]
    assert created["validator_id"] == owner["id"]

    owner_update = api_ok(
        app_client.patch(
            f"/api/validations/{created['id']}",
            headers=owner_headers,
            json={"result": "通过", "description": "本人修订"},
        )
    )["item"]
    assert owner_update["result"] == "通过"
    assert owner_update["description"] == "本人修订"
    assert owner_update["validator_id"] == owner["id"]

    denied = app_client.patch(f"/api/validations/{created['id']}", headers=other_headers, json={"description": "越权修改"})
    assert denied.status_code == 403
    assert "只能编辑自己" in denied.text

    admin_update = api_ok(
        app_client.patch(
            f"/api/validations/{created['id']}",
            headers=auth_headers,
            json={"catalog_no": "VAL-EDIT-ADMIN", "method": "IHC", "description": "管理员修订"},
        )
    )["item"]
    assert admin_update["catalog_no"] == "VAL-EDIT-ADMIN"
    assert admin_update["method"] == "IHC"
    assert admin_update["description"] == "管理员修订"
    assert admin_update["validator_id"] == owner["id"]


def test_history_lists_show_latest_100(app_client, auth_headers):
    import db.database as database

    base_time = datetime(2026, 6, 16, 12, 0, 0)
    with database.connect() as conn:
        user_id = conn.execute("SELECT id FROM users WHERE username = 'admin'").fetchone()["id"]
        for idx in range(101):
            stamp = (base_time - timedelta(minutes=idx)).strftime("%Y-%m-%d %H:%M:%S")
            reagent_id = conn.execute(
                """
                INSERT INTO reagents
                    (code, name, category, catalog_no, quantity, status, storage_node_id, created_by, updated_by, created_at, updated_at)
                VALUES (?, ?, '抗体', ?, 1, '已订购', -2, ?, ?, ?, ?)
                """,
                (f"HIST-ORDER-{idx:03d}", f"历史订购{idx}", f"HIST-ORDER-CAT-{idx:03d}", user_id, user_id, stamp, stamp),
            ).lastrowid
            conn.execute(
                """
                INSERT INTO movements
                    (object_type, object_id, item_type, item_id, from_storage_node_id, to_storage_node_id,
                     from_location_snapshot, to_location_snapshot, moved_by, moved_at, reason, note)
                VALUES (?, ?, 'reagent', ?, -1, -2, '未订购', '未到货', ?, ?, '订购', ?)
                """,
                ("试剂", f"HIST-ORDER-{idx:03d}", reagent_id, user_id, stamp, "history order"),
            )
            conn.execute(
                """
                INSERT INTO validations (catalog_no, validator_id, validation_date, method, result, description, created_at)
                VALUES (?, ?, '2026-06-03', 'WB', '通过', 'bulk validation', ?)
                """,
                (f"HIST-VAL-CAT-{idx:03d}", user_id, stamp),
            )
        conn.commit()

    orders = api_ok(app_client.get("/api/orders", headers=auth_headers))["items"]
    validations = api_ok(app_client.get("/api/validations", headers=auth_headers))["items"]
    movements = api_ok(app_client.get("/api/movements", headers=auth_headers))["items"]

    assert len(orders) == 100
    assert len(validations) == 100
    assert len(movements) == 100
    assert orders[0]["catalog_no"] == "HIST-ORDER-CAT-000"
    assert validations[0]["catalog_no"] == "HIST-VAL-CAT-000"
    assert movements[0]["object_id"] == "HIST-ORDER-000"
    assert "HIST-ORDER-CAT-100" not in {item["catalog_no"] for item in orders}
    assert "HIST-VAL-CAT-100" not in {item["catalog_no"] for item in validations}
    assert "HIST-ORDER-100" not in {item["object_id"] for item in movements}


def test_bulk_import_only_inserts_and_auto_generates_codes(app_client, auth_headers):
    reagent_preview = api_ok(
        app_client.post(
            "/api/bulk/preview",
            headers=auth_headers,
            json={
                "operation": "import",
                "item_type": "reagent",
                "rows": [{"名称": "批量试剂A", "类型": "抗体", "数量": 1, "状态": "可用"}],
            },
        )
    )
    assert reagent_preview["valid"] == 1
    assert reagent_preview["items"][0]["action"] == "新增"

    reagent_commit = api_ok(
        app_client.post(
            "/api/bulk/commit",
            headers=auth_headers,
            json={
                "operation": "import",
                "item_type": "reagent",
                "rows": [{"名称": "批量试剂A", "类型": "抗体", "数量": 1, "状态": "可用"}],
            },
        )
    )
    reagent_item = reagent_commit["items"][0]["item"]
    assert reagent_item["code"].startswith("RG")

    duplicate_preview = api_ok(
        app_client.post(
            "/api/bulk/preview",
            headers=auth_headers,
            json={
                "operation": "import",
                "item_type": "reagent",
                "rows": [{"编号": reagent_item["code"], "名称": "不应更新", "类型": "抗体", "数量": 1, "状态": "可用"}],
            },
        )
    )
    assert duplicate_preview["invalid"] == 1
    assert duplicate_preview["items"][0]["errors"] == ["编号已存在"]

    sample_commit = api_ok(
        app_client.post(
            "/api/bulk/commit",
            headers=auth_headers,
            json={
                "operation": "import",
                "item_type": "sample",
                "rows": [{"样本号": "BULK-SAMPLE-001", "样本类型": "血清", "状态": "可用"}],
            },
        )
    )
    sample_item = sample_commit["items"][0]["item"]
    assert sample_item["code"].startswith("SP")


def test_bulk_validation_and_admin_record_delete_clean_movements(app_client, auth_headers):
    import db.database as database

    reagent = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "Bulk Validation Reagent",
                "category": "抗体",
                "catalog_no": "BULK-VAL-CAT",
                **antibody_fields("CD56", "APC"),
                "quantity": 1,
                "status": "可用",
            },
        ),
        201,
    )["item"]

    validation_preview = api_ok(
        app_client.post(
            "/api/bulk/preview",
            headers=auth_headers,
            json={
                "operation": "validation",
                "rows": [{"货号": "BULK-VAL-CAT", "验证日期": "2026-06-16", "方法": "WB", "结果": "通过", "说明": "批量验证"}],
            },
        )
    )
    assert validation_preview["valid"] == 1
    assert validation_preview["items"][0]["action"] == "验证"

    validation_commit = api_ok(
        app_client.post(
            "/api/bulk/commit",
            headers=auth_headers,
            json={
                "operation": "validation",
                "rows": [{"货号": "BULK-VAL-CAT", "验证日期": "2026-06-16", "方法": "WB", "结果": "通过", "说明": "批量验证"}],
            },
        )
    )
    validation_id = validation_commit["items"][0]["item"]["id"]
    assert validation_commit["success"] == 1

    with database.connect() as conn:
        movement_count = conn.execute(
            "SELECT COUNT(*) AS n FROM movements WHERE item_type = 'reagent' AND item_id = ?",
            (reagent["id"],),
        ).fetchone()["n"]
        assert movement_count >= 1

    deleted_validation = api_ok(
        app_client.post(
            "/api/admin/records/delete",
            headers=auth_headers,
            json={"table": "validations", "ids": [validation_id]},
        )
    )
    assert deleted_validation["count"] == 1

    deleted_reagent = api_ok(
        app_client.post(
            "/api/admin/records/delete",
            headers=auth_headers,
            json={"table": "reagents", "ids": [reagent["id"]]},
        )
    )
    assert deleted_reagent["count"] == 1
    assert deleted_reagent["cleared_refs"]["movement_refs"] == movement_count
    with database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM reagents WHERE id = ?", (reagent["id"],)).fetchone()["n"] == 0
        assert conn.execute(
            "SELECT COUNT(*) AS n FROM movements WHERE item_type = 'reagent' AND item_id = ?",
            (reagent["id"],),
        ).fetchone()["n"] == 0


def test_admin_corrections_update_catalog_no_and_brand_references(app_client, auth_headers):
    import db.database as database
    from services import options_config

    reagent_result = api_ok(
        app_client.post(
            "/api/inventory/items",
            headers=auth_headers,
            json={
                "item_type": "reagent",
                "name": "更正测试试剂",
                "category": "抗体",
                "brand": "Old Co",
                "catalog_no": "CORR-CAT",
                **antibody_fields("CD45", "APC"),
                "quantity": 2,
                "separate_items": True,
                "status": "可用",
            },
        ),
        201,
    )
    items = reagent_result["items"]
    assert len(items) == 2

    api_ok(
        app_client.post(
            "/api/antibody-metadata",
            headers=auth_headers,
            json={
                "catalog_no": "CORR-CAT",
                "target": "CD45",
                "conjugate": "APC",
                "clone": "HI30",
            },
        ),
        201,
    )
    validation_commit = api_ok(
        app_client.post(
            "/api/bulk/commit",
            headers=auth_headers,
            json={
                "operation": "validation",
                "rows": [{"货号": "CORR-CAT", "验证日期": "2026-06-16", "方法": "WB", "结果": "通过"}],
            },
        )
    )
    assert validation_commit["success"] == 1
    catalog_preview = api_ok(
        app_client.post(
            "/api/admin/corrections/catalog-no/preview",
            headers=auth_headers,
            json={"old_value": "CORR-CAT", "new_value": "CORR-CAT-NEW", "reason": "规范化货号"},
        )
    )["item"]
    assert catalog_preview["can_commit"] is True
    assert catalog_preview["counts"]["reagents_catalog_no"] == 2
    assert catalog_preview["counts"]["validations_catalog_no"] == 1
    assert catalog_preview["counts"]["antibody_metadata_catalog_no"] == 1
    catalog_commit = api_ok(
        app_client.post(
            "/api/admin/corrections/catalog-no/commit",
            headers=auth_headers,
            json={"old_value": "CORR-CAT", "new_value": "CORR-CAT-NEW", "reason": "规范化货号"},
        )
    )["item"]
    assert catalog_commit["updated"]["reagents_catalog_no"] == 2
    assert catalog_commit["updated"]["validations_catalog_no"] == 1
    assert catalog_commit["updated"]["antibody_metadata_catalog_no"] == 1
    with database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM reagents WHERE catalog_no = 'CORR-CAT'").fetchone()["n"] == 0
        assert conn.execute("SELECT COUNT(*) AS n FROM validations WHERE catalog_no = 'CORR-CAT-NEW'").fetchone()["n"] == 1
        metadata = conn.execute("SELECT target, clone FROM antibody_metadata WHERE catalog_no = 'CORR-CAT-NEW'").fetchone()
        assert metadata["target"] == "CD45"
        assert metadata["clone"] == "HI30"

    options_config.save_dropdown_options({"brands": ["Old Co", "Existing Co"]})
    brand_preview = api_ok(
        app_client.post(
            "/api/admin/corrections/brand/preview",
            headers=auth_headers,
            json={"old_value": "Old Co", "new_value": "Existing Co", "reason": "统一公司名"},
        )
    )["item"]
    assert brand_preview["can_commit"] is True
    assert brand_preview["counts"]["reagents_brand"] == 2
    assert brand_preview["counts"]["dropdown_brands"] == 1
    assert brand_preview["warnings"]

    brand_commit = api_ok(
        app_client.post(
            "/api/admin/corrections/brand/commit",
            headers=auth_headers,
            json={"old_value": "Old Co", "new_value": "Existing Co", "reason": "统一公司名"},
        )
    )["item"]
    assert brand_commit["updated"]["reagents_brand"] == 2
    assert brand_commit["updated"]["dropdown_brands"] == 1
    with database.connect() as conn:
        assert conn.execute("SELECT COUNT(*) AS n FROM reagents WHERE brand = 'Old Co'").fetchone()["n"] == 0
        assert conn.execute("SELECT COUNT(*) AS n FROM reagents WHERE brand = 'Existing Co'").fetchone()["n"] == 2
    assert options_config.load_dropdown_options()["brands"].count("Existing Co") == 1


def test_validation_image_endpoint_requires_auth_and_serves_file(app_client, auth_headers):
    import routers.registration as registration_routes

    image_dir = registration_routes.VALIDATION_IMAGE_DIR
    image_dir.mkdir(parents=True, exist_ok=True)
    image_path = image_dir / "timeline-preview.png"
    image_path.write_bytes(
        b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?"
        b"\x00\x05\xfe\x02\xfeA\xe2&\xb3\x00\x00\x00\x00IEND\xaeB`\x82"
    )

    app_client.cookies.clear()
    unauthenticated = app_client.get("/api/validation-images/timeline-preview.png")
    assert unauthenticated.status_code == 401

    response = app_client.get("/api/validation-images/timeline-preview.png", headers=auth_headers)
    assert response.status_code == 200
    assert response.content.startswith(b"\x89PNG")

    traversal = app_client.get("/api/validation-images/..%2Fschema.sql", headers=auth_headers)
    assert traversal.status_code in {400, 404}


def test_health_hides_database_path_in_production(app_client, monkeypatch):
    import routers.core as core_routes

    monkeypatch.setattr(core_routes, "IS_PRODUCTION", True)
    production_health = api_ok(app_client.get("/api/health"))
    assert production_health["ok"] is True
    assert production_health["version"] == "1.0.0"
    assert "db" not in production_health

    monkeypatch.setattr(core_routes, "IS_PRODUCTION", False)
    development_health = api_ok(app_client.get("/api/health"))
    assert development_health["version"] == "1.0.0"
    assert "db" in development_health

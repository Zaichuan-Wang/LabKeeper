from __future__ import annotations

import base64
from io import BytesIO
from typing import Any

from core.common import ApiError, create_audit, now_text, row_dict
from core.constants import (
    NODE_TYPE_LABELS,
    PHYSICAL_INVENTORY_STATUS_SQL,
    STATUS_AVAILABLE,
    VALIDATION_UNVERIFIED,
)
from db.database import connect
from services import clinical_samples
from services import movements
from services.options_config import load_dropdown_options
from services import reagents
from services.storage_inventory import get_node, node_full_path, position_options_for_node


REAGENT_TEMPLATE_COLUMNS = [
    "编号", "名称", "类型", "品牌", "货号", "规格量", "规格单位", "数量", "状态", "验证状态",
    "入库日期", "有效期", "存放空间ID", "孔位", "备注",
]
SAMPLE_TEMPLATE_COLUMNS = [
    "系统编号", "样本号", "样本类型", "规格量", "规格单位", "状态", "入库日期", "存放空间ID", "孔位", "备注",
]
CHECKOUT_TEMPLATE_COLUMNS = ["对象类型", "编号", "出库原因", "备注"]
MOVE_TEMPLATE_COLUMNS = ["对象类型", "编号", "目标空间ID", "孔位", "原因", "备注"]
EDIT_TEMPLATE_COLUMNS = [
    "对象类型", "编号", "名称", "类型", "品牌", "货号", "规格量", "规格单位", "数量",
    "状态", "验证状态", "入库日期", "有效期", "存放空间ID", "孔位", "备注",
]
SAMPLE_EDIT_TEMPLATE_COLUMNS = [
    "对象类型", "编号", "样本号", "样本类型", "规格量", "规格单位",
    "状态", "入库日期", "存放空间ID", "孔位", "备注",
]

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_IMPORT_MIMES = {EXCEL_MIME, "application/vnd.ms-excel", "application/octet-stream"}


def _new_workbook() -> Any:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    return Workbook()


def _workbook_bytes(wb: Any) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _xlsx_response(wb: Any, filename: str) -> tuple[bytes, str, str]:
    return _workbook_bytes(wb), EXCEL_MIME, filename


def _set_column_widths(
    ws: Any,
    columns: list[str],
    *,
    minimum: int = 12,
    maximum: int | None = 28,
    extra: int = 8,
    wide_columns: set[str] | None = None,
    wide_width: int = 42,
) -> None:
    try:
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    wide_columns = wide_columns or set()
    for idx, column in enumerate(columns, 1):
        width = wide_width if column in wide_columns else max(minimum, len(column) + extra)
        if maximum is not None:
            width = min(maximum, width)
        ws.column_dimensions[get_column_letter(idx)].width = width


def _item_type(value: Any, default: str = "reagent") -> str:
    text = str(value or default).strip().lower()
    if text in {"sample", "clinical_sample", "clinical-sample", "临床标本", "标本"}:
        return "sample"
    if text in {"reagent", "reagents", "试剂", "耗材", "试剂/耗材"}:
        return "reagent"
    return "reagent"


def _item_type_label(item_type: str) -> str:
    return "临床标本" if item_type == "sample" else "试剂/耗材"


def _parse_excel_data_url(value: str) -> bytes:
    if not value.startswith("data:") or "," not in value:
        raise ApiError(400, "Excel 文件数据格式不正确")
    header, payload = value.split(",", 1)
    mime = header.removeprefix("data:").split(";", 1)[0].lower()
    if mime not in EXCEL_IMPORT_MIMES:
        raise ApiError(400, "只支持 Excel 文件")
    try:
        raw = base64.b64decode(payload)
    except Exception as exc:
        raise ApiError(400, "Excel 文件数据无法解析") from exc
    if len(raw) > 24 * 1024 * 1024:
        raise ApiError(400, "Excel 文件不能超过 24MB")
    return raw


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return value


def parse_excel(data: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    raw = _parse_excel_data_url(str(data.get("data_url", "")))
    sheet_name = str(data.get("sheet", "")).strip()
    wb = load_workbook(BytesIO(raw), data_only=True)
    if sheet_name and sheet_name not in wb.sheetnames:
        raise ApiError(400, "工作表不存在")
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    if ws.max_row < 1:
        raise ApiError(400, "Excel 没有表头")
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    rows: list[dict[str, Any]] = []
    for row_index, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {
            header: _clean_cell(values[idx]) if idx < len(values) else ""
            for idx, header in enumerate(headers)
            if header
        }
        if any(value not in ("", None) for value in record.values()):
            record["_row_no"] = row_index
            rows.append(record)
    return {"items": rows, "count": len(rows), "sheet": ws.title}


def template(query: dict[str, list[str]]) -> tuple[bytes, str, str]:
    operation = query.get("operation", ["import"])[0].strip() or "import"
    item_type = _item_type(query.get("item_type", ["reagent"])[0])
    if operation == "checkout":
        columns = CHECKOUT_TEMPLATE_COLUMNS
        filename = "批量出库模板.xlsx"
    elif operation == "move":
        columns = MOVE_TEMPLATE_COLUMNS
        filename = "批量移动模板.xlsx"
    elif operation == "edit":
        columns = SAMPLE_EDIT_TEMPLATE_COLUMNS if item_type == "sample" else EDIT_TEMPLATE_COLUMNS
        filename = "批量编辑模板.xlsx"
    else:
        columns = SAMPLE_TEMPLATE_COLUMNS if item_type == "sample" else REAGENT_TEMPLATE_COLUMNS
        filename = f"{_item_type_label(item_type)}批量导入模板.xlsx"
    wb = _new_workbook()
    ws = wb.active
    ws.title = "模板"
    ws.append(columns)
    example = _template_example(operation, item_type)
    if example:
        ws.append([example.get(column, "") for column in columns])
    _set_column_widths(ws, columns)
    _append_template_help_sheet(wb, operation, item_type)
    return _xlsx_response(wb, filename)


def _template_example(operation: str, item_type: str) -> dict[str, Any]:
    if operation == "checkout":
        return {"对象类型": "试剂/耗材", "编号": "RG000001", "出库原因": "实验消耗", "备注": "示例行，上传前请删除或改成真实编号"}
    if operation == "move":
        return {"对象类型": "试剂/耗材", "编号": "RG000001", "目标空间ID": 12, "孔位": "A1", "原因": "整理库存", "备注": "目标空间ID请从空间对应表复制"}
    if operation == "edit":
        if item_type == "sample":
            return {
                "对象类型": "临床标本", "编号": "SP000001", "样本号": "P001", "样本类型": "组织",
                "状态": STATUS_AVAILABLE, "存放空间ID": 12, "孔位": "A1",
                "备注": "建议从现有库存清单保留要改的行后编辑",
            }
        return {
            "对象类型": "试剂/耗材", "编号": "RG000001", "名称": "示例抗体", "类型": "抗体",
            "数量": 1, "状态": STATUS_AVAILABLE, "验证状态": VALIDATION_UNVERIFIED, "存放空间ID": 12,
            "孔位": "A1", "备注": "建议从现有库存清单保留要改的行后编辑",
        }
    if item_type == "sample":
        return {
            "样本号": "P001", "样本类型": "血清", "规格量": 0.5, "规格单位": "mL",
            "状态": STATUS_AVAILABLE, "入库日期": "2026-06-12", "存放空间ID": 12, "孔位": "A1",
            "备注": "示例行，上传前请删除或改成真实标本",
        }
    return {
        "编号": "RG000001", "名称": "示例抗体", "类型": "抗体", "品牌": "CST", "货号": "12345",
        "规格量": 100, "规格单位": "uL", "数量": 1, "状态": STATUS_AVAILABLE, "验证状态": VALIDATION_UNVERIFIED,
        "入库日期": "2026-06-12", "有效期": "2027-06-12", "存放空间ID": 12, "孔位": "A1",
        "备注": "示例行，上传前请删除或改成真实试剂",
    }


def _append_template_help_sheet(wb: Any, operation: str, item_type: str) -> None:
    options = load_dropdown_options()
    ws = wb.create_sheet("填写说明")
    ws.append(["项目", "可填内容 / 说明"])
    common_notes = [
        ("对象类型", "批量编辑、移动、出库可填：试剂/耗材、临床标本。单独下载试剂或标本导入模板时不用填对象类型。"),
        ("编号 / 系统编号", "试剂填写试剂编号，例如 RG000001；临床标本的系统编号可在导入时留空自动生成。批量编辑、移动、出库必须填写已有系统编号。"),
        ("样本号", "临床标本的业务样本号，允许重复；例如同一样本号可以分别登记血清和组织。"),
        ("存放空间ID / 目标空间ID", "只填写数字 ID。请点击“下载空间对应表”，从其中复制空间ID；不要填写空间名称或路径。"),
        ("孔位", "只有盒子或带框架的空间需要填写，例如 A1、B3；不需要孔位时留空。"),
        ("空白单元格", "批量编辑时，空白单元格不会覆盖原值；只会修改你填写且与原值不同的字段。"),
    ]
    operation_notes = {
        "import": [("操作流程", "下载模板，删除或改写示例行，填写真实数据，读取 Excel 后先预检查，再确认提交。")],
        "edit": [("操作流程", "推荐先下载现有库存，只保留要编辑的行，修改需要变更的字段；预检查会显示旧值 → 新值。")],
        "move": [("操作流程", "填写对象类型、编号、目标空间ID和孔位；预检查通过后会写入移动记录。")],
        "checkout": [("操作流程", "填写对象类型和编号；确认提交后会出库并释放当前位置。")],
    }
    option_notes = [
        ("试剂类型", "、".join(options.get("categories") or [])),
        ("试剂状态", "、".join(options.get("reagent_statuses") or [])),
        ("验证状态", "、".join(options.get("validation_statuses") or [])),
        ("样本类型", "、".join(options.get("sample_names") or [])),
        ("规格单位", "、".join(options.get("amount_units") or [])),
        ("标本状态", "、".join(options.get("sample_statuses") or [])),
        ("常用品牌", "、".join((options.get("brands") or [])[:20])),
    ]
    for row in operation_notes.get(operation, []) + common_notes + option_notes:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90


def storage_map() -> tuple[bytes, str, str]:
    columns = ["空间ID", "空间名称", "类型", "完整层级位置", "父级ID", "父级名称", "行数", "列数", "可填孔位/格位", "备注"]
    wb = _new_workbook()
    ws = wb.active
    ws.title = "空间对应表"
    ws.append(columns)
    with connect() as conn:
        rows = conn.execute(
            """
            SELECT n.*, p.name AS parent_name
            FROM storage_nodes n LEFT JOIN storage_nodes p ON p.id = n.parent_id
            ORDER BY COALESCE(n.parent_id, 0), n.sort_order, n.id
            """
        ).fetchall()
        for row in rows:
            node = row_dict(row) or {}
            positions = position_options_for_node(row)
            ws.append([
                node.get("id"),
                node.get("name"),
                NODE_TYPE_LABELS.get(str(node.get("node_type")), str(node.get("node_type") or "")),
                node_full_path(conn, int(node["id"])),
                node.get("parent_id") or "",
                node.get("parent_name") or "",
                node.get("rows") or "",
                node.get("cols") or "",
                ", ".join(positions),
                node.get("note") or "",
            ])
    _set_column_widths(ws, columns, maximum=None, extra=6, wide_columns={"完整层级位置", "可填孔位/格位"})
    return _xlsx_response(wb, "空间ID和层级位置对应表.xlsx")


def _current_inventory_row(item: dict[str, Any], item_type: str) -> list[Any]:
    if item_type == "sample":
        return [
            "临床标本", item.get("code") or item.get("id"), item.get("name") or "",
            item.get("category") or "", "", "", item.get("amount"), item.get("amount_unit") or "",
            item.get("quantity") or "", item.get("status") or "", "", item.get("entry_date") or "",
            "", item.get("storage_node_id") or "", item.get("position_in_box") or "", item.get("note") or "",
        ]
    return [
        "试剂/耗材", item.get("code") or item.get("id"), item.get("name") or "",
        item.get("category") or "", item.get("brand") or "", item.get("catalog_no") or "",
        item.get("amount"), item.get("amount_unit") or "", item.get("quantity"),
        item.get("status") or "", item.get("validation_status") or "", item.get("entry_date") or "",
        item.get("expiration_date") or "", item.get("storage_node_id") or "",
        item.get("position_in_box") or "", item.get("note") or "",
    ]


def current_inventory(query: dict[str, list[str]]) -> tuple[bytes, str, str]:
    item_type = str(query.get("item_type", ["all"])[0] or "all").strip()
    if item_type not in {"all", "reagent", "sample"}:
        item_type = "all"
    columns = [
        "对象类型", "编号", "名称", "类别", "品牌", "货号", "规格量", "规格单位",
        "数量", "状态", "验证状态", "入库日期", "有效期", "存放空间ID", "孔位", "备注",
    ]
    wb = _new_workbook()
    ws = wb.active
    ws.title = "现有库存"
    ws.append(columns)
    with connect() as conn:
        if item_type in {"all", "reagent"}:
            rows = conn.execute(
                f"""
                SELECT * FROM reagents
                WHERE COALESCE(status, '') IN {PHYSICAL_INVENTORY_STATUS_SQL}
                ORDER BY updated_at DESC, id DESC
                """
            ).fetchall()
            for row in rows:
                item = row_dict(row) or {}
                ws.append(_current_inventory_row(item, "reagent"))
        if item_type in {"all", "sample"}:
            rows = conn.execute(
                f"""
                SELECT * FROM clinical_samples
                WHERE status IN {PHYSICAL_INVENTORY_STATUS_SQL}
                ORDER BY updated_at DESC, id DESC
                """
            ).fetchall()
            for row in rows:
                item = row_dict(row) or {}
                ws.append(_current_inventory_row(item, "sample"))
    _set_column_widths(ws, columns, maximum=24)
    return _xlsx_response(wb, "现有库存清单.xlsx")


def _resolve_storage_node(conn: Any, row: dict[str, Any], id_key: str = "存放空间ID") -> int | None:
    raw_id = row.get(id_key) or row.get("目标空间ID") or row.get("storage_node_id")
    if raw_id not in ("", None):
        try:
            node_id = int(float(raw_id))
        except (TypeError, ValueError):
            raise ApiError(400, "空间ID必须是数字")
        if node_id <= 0:
            return None
        if get_node(conn, node_id) is None:
            raise ApiError(400, f"空间ID {node_id} 不存在")
        return node_id
    return None


def _find_item(conn: Any, item_type: str, code: Any) -> Any:
    clean = str(code or "").strip()
    if not clean:
        raise ApiError(400, "编号不能为空")
    if item_type == "sample":
        return conn.execute("SELECT * FROM clinical_samples WHERE code = ? ORDER BY id DESC LIMIT 1", (clean,)).fetchone()
    return conn.execute("SELECT * FROM reagents WHERE code = ? ORDER BY id DESC LIMIT 1", (clean,)).fetchone()


def _normalize_import_row(conn: Any, item_type: str, row: dict[str, Any]) -> dict[str, Any]:
    node_id = _resolve_storage_node(conn, row)
    location_requested = any(str(row.get(key) or "").strip() for key in ("存放空间ID", "storage_node_id", "孔位", "position_in_box"))
    if item_type == "sample":
        name = str(row.get("样本号") or row.get("名称") or row.get("name") or "").strip()
        code = str(row.get("系统编号") or row.get("编号") or row.get("code") or "").strip()
        category = str(row.get("样本类型") or row.get("类型") or row.get("category") or "临床标本").strip() or "临床标本"
        if not name:
            raise ApiError(400, "样本号不能为空")
        return {
            "code": code,
            "name": name,
            "category": category,
            "amount": row.get("规格量") or row.get("amount") or None,
            "amount_unit": str(row.get("规格单位") or row.get("amount_unit") or "").strip(),
            "status": str(row.get("状态") or row.get("status") or STATUS_AVAILABLE).strip() or STATUS_AVAILABLE,
            "entry_date": str(row.get("入库日期") or row.get("entry_date") or "").strip(),
            "storage_node_id": node_id,
            "position_in_box": str(row.get("孔位") or row.get("position_in_box") or "").strip(),
            "note": str(row.get("备注") or row.get("note") or "").strip(),
            "_location_requested": location_requested,
        }
    name = str(row.get("名称") or row.get("name") or "").strip()
    if not name:
        raise ApiError(400, "名称不能为空")
    return {
        "code": str(row.get("编号") or row.get("code") or "").strip(),
        "name": name,
        "category": str(row.get("类型") or row.get("category") or "其他").strip() or "其他",
        "brand": str(row.get("品牌") or row.get("brand") or "").strip(),
        "catalog_no": str(row.get("货号") or row.get("catalog_no") or "").strip(),
        "amount": row.get("规格量") or row.get("amount") or None,
        "amount_unit": str(row.get("规格单位") or row.get("单位") or row.get("amount_unit") or "").strip(),
        "quantity": row.get("数量") or row.get("quantity") or 0,
        "status": str(row.get("状态") or row.get("status") or STATUS_AVAILABLE).strip() or STATUS_AVAILABLE,
        "validation_status": str(row.get("验证状态") or row.get("validation_status") or VALIDATION_UNVERIFIED).strip() or VALIDATION_UNVERIFIED,
        "entry_date": str(row.get("入库日期") or row.get("entry_date") or "").strip(),
        "expiration_date": str(row.get("有效期") or row.get("expiration_date") or "").strip(),
        "storage_node_id": node_id,
        "position_in_box": str(row.get("孔位") or row.get("position_in_box") or "").strip(),
        "note": str(row.get("备注") or row.get("note") or "").strip(),
        "_location_requested": location_requested,
    }


def _normalize_item_ref(conn: Any, row: dict[str, Any], default_type: str) -> tuple[str, Any]:
    item_type = _item_type(row.get("对象类型") or row.get("item_type") or default_type, default_type)
    code = row.get("编号") or row.get("code")
    item = _find_item(conn, item_type, code)
    if item is None:
        raise ApiError(404, f"{_item_type_label(item_type)}不存在：{code}")
    return item_type, item


FIELD_LABELS = {
    "name": "名称",
    "category": "类型",
    "brand": "品牌",
    "catalog_no": "货号",
    "amount": "规格量",
    "amount_unit": "规格单位",
    "quantity": "数量",
    "status": "状态",
    "validation_status": "验证状态",
    "entry_date": "入库日期",
    "expiration_date": "有效期",
    "storage_node_id": "存放空间ID",
    "position_in_box": "孔位",
    "note": "备注",
}


def _same_value(old: Any, new: Any) -> bool:
    old_text = "" if old is None else str(old).strip()
    new_text = "" if new is None else str(new).strip()
    if old_text == new_text:
        return True
    try:
        return float(old_text) == float(new_text)
    except (TypeError, ValueError):
        return False


def _normalize_edit_row(conn: Any, item_type: str, row: dict[str, Any]) -> dict[str, Any]:
    ref_type, item = _normalize_item_ref(conn, row, item_type)
    updates: dict[str, Any] = {"item_type": ref_type, "item_id": int(item["id"])}
    changes: list[str] = []
    if ref_type == "sample":
        mapping = {
            "name": row.get("样本号") or row.get("名称") or row.get("name"),
            "category": row.get("样本类型") or row.get("类型") or row.get("category"),
            "amount": row.get("规格量") or row.get("amount"),
            "amount_unit": row.get("规格单位") or row.get("amount_unit"),
            "status": row.get("状态") or row.get("status"),
            "entry_date": row.get("入库日期") or row.get("entry_date"),
            "note": row.get("备注") or row.get("note"),
        }
    else:
        mapping = {
            "name": row.get("名称") or row.get("name"),
            "category": row.get("类型") or row.get("category"),
            "brand": row.get("品牌") or row.get("brand"),
            "catalog_no": row.get("货号") or row.get("catalog_no"),
            "amount": row.get("规格量") or row.get("amount"),
            "amount_unit": row.get("规格单位") or row.get("amount_unit"),
            "quantity": row.get("数量") or row.get("quantity"),
            "status": row.get("状态") or row.get("status"),
            "validation_status": row.get("验证状态") or row.get("validation_status"),
            "entry_date": row.get("入库日期") or row.get("entry_date"),
            "expiration_date": row.get("有效期") or row.get("expiration_date"),
            "note": row.get("备注") or row.get("note"),
        }
    for key, value in mapping.items():
        if value not in ("", None) and not _same_value(item[key], value):
            updates[key] = value
            changes.append(f"{FIELD_LABELS.get(key, key)}：{item[key] if item[key] not in (None, '') else '空'} → {value}")
    has_location = any(str(row.get(key) or "").strip() for key in ("存放空间ID", "storage_node_id", "孔位", "position_in_box"))
    if has_location:
        node_id = _resolve_storage_node(conn, row)
        position = str(row.get("孔位") or row.get("position_in_box") or "").strip()
        if not _same_value(item["storage_node_id"], node_id):
            updates["storage_node_id"] = node_id
            changes.append(f"存放空间ID：{item['storage_node_id'] if item['storage_node_id'] not in (None, '') else '空'} → {node_id if node_id is not None else '空'}")
        if not _same_value(item["position_in_box"], position):
            updates["position_in_box"] = position
            changes.append(f"孔位：{item['position_in_box'] if item['position_in_box'] not in (None, '') else '空'} → {position or '空'}")
    if len(updates) <= 2:
        raise ApiError(400, "没有字段发生变化")
    updates["_changes"] = changes
    return updates


def _preview_one(conn: Any, operation: str, item_type: str, mode: str, row: dict[str, Any]) -> dict[str, Any]:
    if operation == "import":
        normalized = _normalize_import_row(conn, item_type, row)
        code_key = "code"
        code = normalized.get(code_key)
        existing = _find_item(conn, item_type, code) if code else None
        if mode == "update" and not code:
            raise ApiError(400, "更新已有临床标本时必须填写系统编号")
        if mode == "insert" and existing:
            raise ApiError(409, "编号已存在")
        if mode == "update" and not existing:
            raise ApiError(404, "编号不存在，不能更新")
        action = "更新" if existing else "新增"
        return {"action": action, "normalized": normalized}
    if operation == "move":
        ref_type, item = _normalize_item_ref(conn, row, item_type)
        node_id = _resolve_storage_node(conn, row, "目标空间ID")
        return {"action": "移动", "normalized": {"item_type": ref_type, "item_id": item["id"], "to_storage_node_id": node_id, "position_in_box": str(row.get("孔位") or "").strip()}}
    if operation == "edit":
        normalized = _normalize_edit_row(conn, item_type, row)
        return {"action": "编辑", "normalized": normalized, "summary": "；".join(normalized.get("_changes") or [])}
    if operation == "checkout":
        ref_type, item = _normalize_item_ref(conn, row, item_type)
        return {"action": "出库", "normalized": {"item_type": ref_type, "item_id": item["id"]}}
    raise ApiError(400, "批量操作类型不正确")


def preview(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    operation = str(data.get("operation", "import")).strip() or "import"
    item_type = _item_type(data.get("item_type", "reagent"))
    mode = str(data.get("mode", "upsert")).strip() or "upsert"
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        raise ApiError(400, "批量数据格式不正确")
    items = []
    with connect() as conn:
        for index, row in enumerate(rows, start=1):
            row_no = row.get("_row_no") or index
            try:
                result = _preview_one(conn, operation, item_type, mode, row)
                items.append({"row_no": row_no, "status": "ok", "errors": [], "summary": result.get("summary", ""), **result, "source": row})
            except ApiError as exc:
                items.append({"row_no": row_no, "status": "error", "errors": [exc.message], "action": "", "normalized": {}, "source": row})
    valid = sum(1 for item in items if item["status"] == "ok")
    return {"items": items, "total": len(items), "valid": valid, "invalid": len(items) - valid}


def _commit_import(item_type: str, mode: str, normalized: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    payload = dict(normalized)
    location_requested = bool(payload.pop("_location_requested", False))
    with connect() as conn:
        code_key = "code"
        code = payload.get(code_key)
        existing = _find_item(conn, item_type, code) if code else None
    if existing:
        if mode == "insert":
            raise ApiError(409, "编号已存在")
        if not location_requested:
            payload.pop("storage_node_id", None)
            payload.pop("position_in_box", None)
        if item_type == "sample":
            return clinical_samples.update_sample(int(existing["id"]), payload, user)["item"]
        return reagents.update_reagent(int(existing["id"]), payload, user)["item"]
    if mode == "update":
        raise ApiError(404, "编号不存在，不能更新")
    if item_type == "sample":
        return clinical_samples.create_sample(payload, user)["item"]
    return reagents.create_reagent(payload, user)["item"]


def commit(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    operation = str(data.get("operation", "import")).strip() or "import"
    item_type = _item_type(data.get("item_type", "reagent"))
    mode = str(data.get("mode", "upsert")).strip() or "upsert"
    rows = data.get("rows") or []
    if not isinstance(rows, list):
        raise ApiError(400, "批量数据格式不正确")
    timestamp = now_text()
    results = []
    success = 0
    failed = 0
    for index, row in enumerate(rows, start=1):
        row_no = row.get("_row_no") or index
        try:
            with connect() as conn:
                preview_result = _preview_one(conn, operation, item_type, mode, row)
            normalized = preview_result["normalized"]
            if operation == "import":
                item = _commit_import(item_type, mode, normalized, user)
            elif operation == "edit":
                payload = dict(normalized)
                payload.pop("_changes", None)
                ref_type = payload.pop("item_type")
                item_id = int(payload.pop("item_id"))
                if ref_type == "sample":
                    item = clinical_samples.update_sample(item_id, payload, user)["item"]
                else:
                    item = reagents.update_reagent(item_id, payload, user)["item"]
            elif operation == "move":
                payload = {
                    **normalized,
                    "reason": str(row.get("原因") or row.get("reason") or "批量移动").strip() or "批量移动",
                    "note": str(row.get("备注") or row.get("note") or "").strip(),
                }
                item = movements.create_movement(payload, user)["item"]
            elif operation == "checkout":
                payload = {
                    **normalized,
                    "reason": str(row.get("出库原因") or row.get("reason") or "批量出库").strip() or "批量出库",
                    "note": str(row.get("备注") or row.get("note") or "").strip(),
                }
                item = movements.create_checkout(payload, user)["item"]
            else:
                raise ApiError(400, "批量操作类型不正确")
            success += 1
            results.append({"row_no": row_no, "status": "ok", "action": preview_result["action"], "item": item, "errors": []})
        except ApiError as exc:
            failed += 1
            results.append({"row_no": row_no, "status": "error", "action": "", "item": None, "errors": [exc.message]})
    with connect() as conn:
        create_audit(conn, user["id"], "api_bulk_operation", operation, None, {
            "item_type": item_type, "mode": mode, "total": len(rows), "success": success, "failed": failed, "created_at": timestamp,
        })
        conn.commit()
    return {"items": results, "total": len(rows), "success": success, "failed": failed}

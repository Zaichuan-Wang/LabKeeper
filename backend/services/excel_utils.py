from __future__ import annotations

import base64
from io import BytesIO
from typing import Any

from core.common import ApiError


EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_IMPORT_EXTENSIONS = {
    EXCEL_MIME: ".xlsx",
    "application/vnd.ms-excel": ".xls",
    "application/octet-stream": ".xlsx",
}
MAX_EXCEL_BYTES = 24 * 1024 * 1024


def new_workbook() -> Any:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    return Workbook()


def workbook_bytes(wb: Any) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def xlsx_response(wb: Any, filename: str) -> tuple[bytes, str, str]:
    return workbook_bytes(wb), EXCEL_MIME, filename


def parse_excel_data_url(value: str) -> tuple[bytes, str]:
    if not value.startswith("data:") or "," not in value:
        raise ApiError(400, "Excel 文件数据格式不正确")
    header, payload = value.split(",", 1)
    mime = header.removeprefix("data:").split(";", 1)[0].lower()
    if mime not in EXCEL_IMPORT_EXTENSIONS:
        raise ApiError(400, "只支持 Excel 文件")
    try:
        body = base64.b64decode(payload)
    except Exception as exc:
        raise ApiError(400, "Excel 文件数据无法解析") from exc
    if len(body) > MAX_EXCEL_BYTES:
        raise ApiError(400, "Excel 文件不能超过 24MB")
    return body, EXCEL_IMPORT_EXTENSIONS[mime]


def clean_excel_cell(value: Any, *, blank_as_none: bool = False) -> Any:
    if value is None:
        return None if blank_as_none else ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, str):
        text = value.strip()
        return None if blank_as_none and text == "" else text
    return value


def excel_export_cell(value: Any) -> Any:
    if isinstance(value, bytes):
        return base64.b64encode(value).decode("ascii")
    return value


def set_column_widths(
    ws: Any,
    columns: list[str],
    *,
    minimum: int = 12,
    maximum: int | None = 28,
    extra: int = 8,
    wide_columns: set[str] | None = None,
    wide_width: int = 42,
) -> None:
    try:
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    wide_columns = wide_columns or set()
    for idx, column in enumerate(columns, 1):
        width = wide_width if column in wide_columns else max(minimum, len(column) + extra)
        if maximum is not None:
            width = min(maximum, width)
        ws.column_dimensions[get_column_letter(idx)].width = width

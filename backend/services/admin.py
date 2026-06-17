from __future__ import annotations

import json
from io import BytesIO
from typing import Any

from services import backup as database_backup
from core import config
from core.common import ApiError, clean_int_range, create_audit, now_text, row_dict
from core.constants import DEFAULT_USER_PERMISSIONS, PERMISSIONS, ROLES
from db.database import connect
from services.auth import hash_password, user_permissions
from services.excel_utils import clean_excel_cell, excel_export_cell, parse_excel_data_url
from services.options_config import load_dropdown_options, save_dropdown_options


def options() -> dict[str, Any]:
    dropdowns = load_dropdown_options()
    return {
        **dropdowns,
        "roles": ROLES,
        "permissions": PERMISSIONS,
        "default_user_permissions": DEFAULT_USER_PERMISSIONS,
    }


def dropdown_options() -> dict[str, Any]:
    return {"item": load_dropdown_options()}


def update_dropdown_options(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    old_options = load_dropdown_options()
    clean = save_dropdown_options(data)
    with connect() as conn:
        create_audit(conn, user["id"], "api_update_dropdown_options", "dropdown_options", None, clean, old_options)
        conn.commit()
    return {"item": clean}


def users() -> dict[str, Any]:
    with connect() as conn:
        rows = conn.execute("SELECT id, username, display_name, role, permissions, is_active, created_at, updated_at FROM users ORDER BY id").fetchall()
    items = [_public_user_row(row) for row in rows]
    return {"items": items, "count": len(items)}


def create_user(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    username = str(data.get("username", "")).strip()
    password = str(data.get("password", ""))
    display_name = str(data.get("display_name", "")).strip()
    role = _clean_role(data.get("role"))
    if role not in ROLES:
        raise ApiError(400, "角色不正确")
    if not username:
        raise ApiError(400, "用户名不能为空")
    if not password:
        password = _initial_password()
    permissions = _permissions_json(data.get("permissions"), role)
    timestamp = now_text()
    with connect() as conn:
        cur = conn.execute(
            """
            INSERT INTO users (username, display_name, password_hash, role, permissions, is_active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?)
            """,
            (username, display_name, hash_password(password), role, permissions, timestamp, timestamp),
        )
        create_audit(conn, user["id"], "api_create_user", "users", cur.lastrowid, {"username": username, "role": role})
        conn.commit()
        row = conn.execute("SELECT id, username, display_name, role, permissions, is_active, created_at, updated_at FROM users WHERE id = ?", (cur.lastrowid,)).fetchone()
    return {"item": _public_user_row(row)}


def update_user(target_id: int, data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    allowed = {"display_name", "role", "is_active", "permissions"}
    updates = {key: data[key] for key in allowed if key in data}
    if "role" in updates:
        updates["role"] = _clean_role(updates["role"])
        if updates["role"] not in ROLES:
            raise ApiError(400, "角色不正确")
    if "is_active" in updates:
        updates["is_active"] = 1 if bool(updates["is_active"]) else 0
    if "permissions" in updates:
        role_for_permissions = str(updates.get("role") or "")
        if not role_for_permissions:
            with connect() as conn:
                old_role = conn.execute("SELECT role FROM users WHERE id = ?", (target_id,)).fetchone()
            role_for_permissions = old_role["role"] if old_role else "user"
        updates["permissions"] = _permissions_json(updates["permissions"], role_for_permissions)
    if data.get("password"):
        updates["password_hash"] = hash_password(str(data["password"]))
    if not updates:
        raise ApiError(400, "没有可更新字段")
    updates["updated_at"] = now_text()
    assignments = ", ".join(f"{key} = ?" for key in updates)
    with connect() as conn:
        old = conn.execute("SELECT id, username, display_name, role, permissions, is_active FROM users WHERE id = ?", (target_id,)).fetchone()
        if old is None:
            raise ApiError(404, "用户不存在")
        next_role = str(updates.get("role", old["role"]))
        next_active = int(updates.get("is_active", old["is_active"]))
        _ensure_not_last_enabled_admin(conn, target_id, next_role, next_active)
        conn.execute(f"UPDATE users SET {assignments} WHERE id = ?", list(updates.values()) + [target_id])
        create_audit(conn, user["id"], "api_update_user", "users", target_id, data, row_dict(old))
        conn.commit()
        row = conn.execute("SELECT id, username, display_name, role, permissions, is_active, created_at, updated_at FROM users WHERE id = ?", (target_id,)).fetchone()
    return {"item": _public_user_row(row)}


def reset_user_password(target_id: int, user: dict[str, Any]) -> dict[str, Any]:
    password = _initial_password()
    timestamp = now_text()
    with connect() as conn:
        old = conn.execute("SELECT id, username, display_name, role, permissions, is_active FROM users WHERE id = ?", (target_id,)).fetchone()
        if old is None:
            raise ApiError(404, "用户不存在")
        conn.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?",
            (hash_password(password), timestamp, target_id),
        )
        create_audit(conn, user["id"], "api_reset_user_password", "users", target_id, {"username": old["username"]})
        conn.commit()
        row = conn.execute("SELECT id, username, display_name, role, permissions, is_active, created_at, updated_at FROM users WHERE id = ?", (target_id,)).fetchone()
    return {"item": _public_user_row(row)}


def _public_user_row(row: Any) -> dict[str, Any]:
    item = row_dict(row) or {}
    item["role"] = _clean_role(item.get("role"))
    item["permissions"] = user_permissions(item.get("permissions"), item["role"])
    return item


def _initial_password() -> str:
    password = str(config.INITIAL_PASSWORD or "")
    if not password:
        raise ApiError(500, "系统初始密码未配置")
    return password


def _clean_role(value: Any) -> str:
    role = str(value or "user").strip()
    return role if role in ROLES else "user"


def _permissions_json(raw: Any, role: str) -> str:
    permissions = user_permissions(raw, role)
    if role == "admin":
        permissions = {}
    return json.dumps(permissions, ensure_ascii=False, separators=(",", ":"))


def _ensure_not_last_enabled_admin(conn: Any, target_id: int, next_role: str, next_active: int) -> None:
    old = conn.execute("SELECT role, is_active FROM users WHERE id = ?", (target_id,)).fetchone()
    if old is None or old["role"] != "admin" or int(old["is_active"]) != 1:
        return
    if next_role == "admin" and int(next_active) == 1:
        return
    others = conn.execute(
        "SELECT COUNT(*) AS n FROM users WHERE id != ? AND role = 'admin' AND is_active = 1",
        (target_id,),
    ).fetchone()["n"]
    if int(others or 0) <= 0:
        raise ApiError(400, "不能停用或降级最后一个启用管理员")


def _ensure_enabled_admin_exists(conn: Any) -> None:
    row = conn.execute("SELECT COUNT(*) AS n FROM users WHERE role = 'admin' AND is_active = 1").fetchone()
    if int(row["n"] or 0) <= 0:
        raise ApiError(400, "导入后没有启用状态的管理员，已取消写入")


def excel_tables() -> dict[str, Any]:
    with connect() as conn:
        names = _table_names(conn)
        items = [{"name": name} for name in names]
    return {"items": items, "count": len(items)}


def excel_export(query: dict[str, list[str]]) -> tuple[bytes, str, str]:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    table = query.get("table", [""])[0].strip()
    mode = query.get("mode", ["data"])[0].strip() or "data"
    limit = clean_int_range(query.get("limit", ["0"])[0], 0, 0, 100_000)
    with connect() as conn:
        names = _table_names(conn)
        if table and table not in names:
            raise ApiError(400, "数据表不存在")
        export_names = [table] if table else names
        wb = Workbook()
        wb.remove(wb.active)
        for table_name in export_names:
            ws = wb.create_sheet(title=table_name[:31])
            columns = _table_columns(conn, table_name)
            ws.append(columns)
            if mode != "template":
                sql = f'SELECT * FROM "{table_name}"'
                if "updated_at" in columns:
                    sql += " ORDER BY updated_at DESC"
                elif "created_at" in columns:
                    sql += " ORDER BY created_at DESC"
                elif "id" in columns:
                    sql += " ORDER BY id"
                rows = conn.execute(sql + " LIMIT ?", (limit,)).fetchall() if limit > 0 else conn.execute(sql).fetchall()
                for row in rows:
                    ws.append([excel_export_cell(row[col]) for col in columns])
            for idx, column in enumerate(columns, 1):
                ws.column_dimensions[get_column_letter(idx)].width = max(12, min(32, len(column) + 4))
    buffer = BytesIO()
    wb.save(buffer)
    suffix = "template" if mode == "template" else "export"
    filename = f"{table or 'all_tables'}_{suffix}.xlsx"
    return buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename


def excel_import(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ApiError(500, "Excel 功能需要安装 openpyxl") from exc
    file_data = str(data.get("data_url", ""))
    mode = str(data.get("mode", "append")).strip()
    scope = str(data.get("scope", "single")).strip()
    target_table = str(data.get("table", "")).strip()
    sheet_name = str(data.get("sheet", "")).strip()
    if mode not in {"append", "upsert"}:
        raise ApiError(400, "导入模式不正确")
    raw, _ = parse_excel_data_url(file_data)
    wb = load_workbook(BytesIO(raw), data_only=True)
    summary: list[dict[str, Any]] = []
    with connect() as conn:
        names = _table_names(conn)
        if scope == "workbook":
            jobs = [(name, name) for name in wb.sheetnames if name in names]
            if not jobs:
                raise ApiError(400, "没有可导入的工作表")
        else:
            if target_table not in names:
                raise ApiError(400, "目标表不存在")
            selected_sheet = sheet_name or wb.sheetnames[0]
            if selected_sheet not in wb.sheetnames:
                raise ApiError(400, "工作表不存在")
            jobs = [(target_table, selected_sheet)]
    backup_item = database_backup.create_database_backup("before_excel_import", user)
    with connect() as conn:
        for table_name, sheet in jobs:
            result = _import_sheet(conn, table_name, wb[sheet], mode)
            summary.append({"table": table_name, "sheet": sheet, **result, "failed": len(result["failed"])})
            if result["failed"]:
                summary[-1]["failed_rows"] = result["failed"][:50]
            create_audit(conn, user["id"], "api_import_excel", table_name, None, {"mode": mode, "sheet": sheet, **summary[-1]})
        if any(table_name == "users" for table_name, _ in jobs):
            _ensure_enabled_admin_exists(conn)
        conn.commit()
    return {"items": summary, "backup": backup_item}


DELETE_RECORD_TABLES = {
    "reagents": "试剂记录",
    "clinical_samples": "临床标本记录",
    "validations": "验证记录",
    "movements": "移动记录",
}


def delete_records(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    table = str(data.get("table") or "").strip()
    if table not in DELETE_RECORD_TABLES:
        raise ApiError(400, "这张表不支持在页面直接删除记录")
    ids = _clean_delete_ids(data.get("ids"))
    if not ids:
        raise ApiError(400, "请填写要删除的记录 ID")
    backup_item = database_backup.create_database_backup(f"before_delete_{table}", user)
    placeholders = ", ".join("?" for _ in ids)
    with connect() as conn:
        existing = conn.execute(
            f"SELECT * FROM {_quote_ident(table)} WHERE id IN ({placeholders}) ORDER BY id",
            ids,
        ).fetchall()
        if len(existing) != len(ids):
            found = {int(row["id"]) for row in existing}
            missing = [item_id for item_id in ids if item_id not in found]
            raise ApiError(404, f"记录不存在：{', '.join(str(item_id) for item_id in missing)}")
        old_rows = [row_dict(row) or {} for row in existing]
        cleared_refs = _clear_delete_references(conn, table, ids)
        conn.execute(f"DELETE FROM {_quote_ident(table)} WHERE id IN ({placeholders})", ids)
        create_audit(conn, user["id"], "api_delete_records", table, None, {"ids": ids, "count": len(ids), "cleared_refs": cleared_refs}, old_rows)
        conn.commit()
    return {
        "table": table,
        "label": DELETE_RECORD_TABLES[table],
        "ids": ids,
        "count": len(ids),
        "items": old_rows,
        "cleared_refs": cleared_refs,
        "backup": backup_item,
    }


def _clean_correction_values(data: dict[str, Any], label: str) -> tuple[str, str, str]:
    old_value = str(data.get("old_value") or "").strip()
    new_value = str(data.get("new_value") or "").strip()
    reason = str(data.get("reason") or "").strip()
    if not old_value:
        raise ApiError(400, f"原{label}不能为空")
    if not new_value:
        raise ApiError(400, f"新{label}不能为空")
    if old_value == new_value:
        raise ApiError(400, f"原{label}和新{label}相同，无需更正")
    return old_value, new_value, reason


def _count_where(conn: Any, sql: str, params: list[Any] | tuple[Any, ...]) -> int:
    row = conn.execute(sql, params).fetchone()
    return int(row["n"] or 0)


def _catalog_no_examples(conn: Any, old_catalog: str) -> list[dict[str, Any]]:
    reagent_rows = conn.execute(
        """
        SELECT 'reagents' AS source_table, id, code, source_code, name, brand, catalog_no, status
        FROM reagents
        WHERE TRIM(COALESCE(catalog_no, '')) = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT 8
        """,
        (old_catalog,),
    ).fetchall()
    metadata_rows = conn.execute(
        """
        SELECT 'antibody_metadata' AS source_table, catalog_no, target, conjugate, clone
        FROM antibody_metadata
        WHERE TRIM(COALESCE(catalog_no, '')) = ?
        ORDER BY updated_at DESC, catalog_no
        LIMIT 4
        """,
        (old_catalog,),
    ).fetchall()
    validation_rows = conn.execute(
        """
        SELECT 'validations' AS source_table, id, catalog_no, validation_date, method, result
        FROM validations
        WHERE TRIM(COALESCE(catalog_no, '')) = ?
        ORDER BY created_at DESC, id DESC
        LIMIT 4
        """,
        (old_catalog,),
    ).fetchall()
    return [row_dict(row) or {} for row in [*reagent_rows, *metadata_rows, *validation_rows]][:10]


def preview_catalog_no_correction(data: dict[str, Any]) -> dict[str, Any]:
    old_catalog, new_catalog, reason = _clean_correction_values(data, "货号")
    with connect() as conn:
        reagent_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM reagents WHERE TRIM(COALESCE(catalog_no, '')) = ?", (old_catalog,))
        validation_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM validations WHERE TRIM(COALESCE(catalog_no, '')) = ?", (old_catalog,))
        metadata_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM antibody_metadata WHERE TRIM(COALESCE(catalog_no, '')) = ?", (old_catalog,))
        target_metadata_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM antibody_metadata WHERE TRIM(COALESCE(catalog_no, '')) = ?", (new_catalog,))
        examples = _catalog_no_examples(conn, old_catalog)
    total = reagent_rows + validation_rows + metadata_rows
    blocking: list[str] = []
    warnings: list[str] = []
    if total <= 0:
        blocking.append("没有找到这个旧货号的业务引用或抗体元信息")
    if target_metadata_rows:
        blocking.append("新货号已存在抗体元信息，不能直接改写主键；请先合并或删除重复元信息")
    if validation_rows:
        warnings.append("验证记录按货号关联，提交后会同步迁移到新货号")
    return {
        "item": {
            "kind": "catalog_no",
            "old_value": old_catalog,
            "new_value": new_catalog,
            "reason": reason,
            "can_commit": not blocking,
            "blocking": blocking,
            "warnings": warnings,
            "counts": {
                "reagents_catalog_no": reagent_rows,
                "validations_catalog_no": validation_rows,
                "antibody_metadata_catalog_no": metadata_rows,
                "target_antibody_metadata_catalog_no": target_metadata_rows,
                "total": total,
            },
            "examples": examples,
        }
    }


def commit_catalog_no_correction(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    preview = preview_catalog_no_correction(data)["item"]
    if not preview["can_commit"]:
        raise ApiError(409, "预检查未通过：" + "；".join(preview["blocking"]))
    old_catalog = preview["old_value"]
    new_catalog = preview["new_value"]
    reason = preview["reason"]
    backup_item = database_backup.create_database_backup("before_catalog_no_correction", user)
    timestamp = now_text()
    with connect() as conn:
        examples = _catalog_no_examples(conn, old_catalog)
        updated: dict[str, int] = {}
        cur = conn.execute(
            "UPDATE reagents SET catalog_no = ?, updated_by = ?, updated_at = ? WHERE TRIM(COALESCE(catalog_no, '')) = ?",
            (new_catalog, user["id"], timestamp, old_catalog),
        )
        updated["reagents_catalog_no"] = int(cur.rowcount or 0)
        cur = conn.execute(
            "UPDATE validations SET catalog_no = ? WHERE TRIM(COALESCE(catalog_no, '')) = ?",
            (new_catalog, old_catalog),
        )
        updated["validations_catalog_no"] = int(cur.rowcount or 0)
        cur = conn.execute(
            "UPDATE antibody_metadata SET catalog_no = ?, updated_by = ?, updated_at = ? WHERE TRIM(COALESCE(catalog_no, '')) = ?",
            (new_catalog, user["id"], timestamp, old_catalog),
        )
        updated["antibody_metadata_catalog_no"] = int(cur.rowcount or 0)
        create_audit(
            conn,
            user["id"],
            "api_correct_catalog_no",
            "reagents",
            None,
            {
                "old_catalog_no": old_catalog,
                "new_catalog_no": new_catalog,
                "reason": reason,
                "updated": updated,
                "backup": backup_item.get("filename"),
            },
            {"examples": examples},
        )
        conn.commit()
    result = dict(preview)
    result["updated"] = updated
    result["backup"] = backup_item
    return {"item": result}


def _brand_examples(conn: Any, old_brand: str) -> list[dict[str, Any]]:
    rows = conn.execute(
        """
        SELECT 'reagents' AS source_table, id, code, source_code, name, brand, catalog_no, status
        FROM reagents
        WHERE TRIM(COALESCE(brand, '')) = ?
        ORDER BY updated_at DESC, id DESC
        LIMIT 10
        """,
        (old_brand,),
    ).fetchall()
    return [row_dict(row) or {} for row in rows]


def preview_brand_correction(data: dict[str, Any]) -> dict[str, Any]:
    old_brand, new_brand, reason = _clean_correction_values(data, "公司名")
    options = load_dropdown_options()
    brands = options.get("brands") or []
    with connect() as conn:
        reagent_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM reagents WHERE TRIM(COALESCE(brand, '')) = ?", (old_brand,))
        target_rows = _count_where(conn, "SELECT COUNT(*) AS n FROM reagents WHERE TRIM(COALESCE(brand, '')) = ?", (new_brand,))
        examples = _brand_examples(conn, old_brand)
    option_rows = sum(1 for brand in brands if str(brand).strip() == old_brand)
    target_option_rows = sum(1 for brand in brands if str(brand).strip() == new_brand)
    total = reagent_rows + option_rows
    blocking: list[str] = []
    warnings: list[str] = []
    if total <= 0:
        blocking.append("没有找到这个旧公司名的业务引用或常用品牌选项")
    if target_rows or target_option_rows:
        warnings.append("新公司名已经存在，提交后会把旧名称合并到现有名称")
    return {
        "item": {
            "kind": "brand",
            "old_value": old_brand,
            "new_value": new_brand,
            "reason": reason,
            "can_commit": not blocking,
            "blocking": blocking,
            "warnings": warnings,
            "counts": {
                "reagents_brand": reagent_rows,
                "dropdown_brands": option_rows,
                "target_reagents_brand": target_rows,
                "target_dropdown_brands": target_option_rows,
                "total": total,
            },
            "examples": examples,
        }
    }


def _correct_brand_options(old_brand: str, new_brand: str) -> tuple[dict[str, Any], dict[str, Any], int]:
    old_options = load_dropdown_options()
    old_brands = old_options.get("brands") or []
    changed = 0
    next_brands: list[str] = []
    for brand in old_brands:
        clean = str(brand).strip()
        next_value = new_brand if clean == old_brand else clean
        if clean == old_brand:
            changed += 1
        if next_value and next_value not in next_brands:
            next_brands.append(next_value)
    if changed <= 0:
        return old_options, old_options, 0
    next_options = {**old_options, "brands": next_brands}
    return old_options, save_dropdown_options(next_options), changed


def commit_brand_correction(data: dict[str, Any], user: dict[str, Any]) -> dict[str, Any]:
    preview = preview_brand_correction(data)["item"]
    if not preview["can_commit"]:
        raise ApiError(409, "预检查未通过：" + "；".join(preview["blocking"]))
    old_brand = preview["old_value"]
    new_brand = preview["new_value"]
    reason = preview["reason"]
    backup_item = database_backup.create_database_backup("before_brand_correction", user)
    timestamp = now_text()
    with connect() as conn:
        examples = _brand_examples(conn, old_brand)
        cur = conn.execute(
            "UPDATE reagents SET brand = ?, updated_by = ?, updated_at = ? WHERE TRIM(COALESCE(brand, '')) = ?",
            (new_brand, user["id"], timestamp, old_brand),
        )
        reagent_updates = int(cur.rowcount or 0)
        old_options, new_options, option_updates = _correct_brand_options(old_brand, new_brand)
        updated = {"reagents_brand": reagent_updates, "dropdown_brands": option_updates}
        create_audit(
            conn,
            user["id"],
            "api_correct_brand",
            "reagents",
            None,
            {
                "old_brand": old_brand,
                "new_brand": new_brand,
                "reason": reason,
                "updated": updated,
                "backup": backup_item.get("filename"),
            },
            {"examples": examples, "dropdown_options": old_options if option_updates else None},
        )
        if option_updates:
            create_audit(conn, user["id"], "api_update_dropdown_options", "dropdown_options", None, new_options, old_options)
        conn.commit()
    result = dict(preview)
    result["updated"] = updated
    result["backup"] = backup_item
    return {"item": result}


def _clear_delete_references(conn: Any, table: str, ids: list[int]) -> dict[str, int]:
    placeholders = ", ".join("?" for _ in ids)
    if table in {"reagents", "clinical_samples"}:
        item_type = "reagent" if table == "reagents" else "sample"
        movement_rows = conn.execute(
            f"SELECT id FROM movements WHERE item_type = ? AND item_id IN ({placeholders})",
            [item_type, *ids],
        ).fetchall()
        movement_ids = [int(row["id"]) for row in movement_rows]
        rollback_refs = _clear_movement_rollback_references(conn, movement_ids) if movement_ids else 0
        conn.execute(
            f"DELETE FROM movements WHERE item_type = ? AND item_id IN ({placeholders})",
            [item_type, *ids],
        )
        return {"movement_refs": len(movement_ids), "movement_rollback_refs": rollback_refs}
    if table != "movements":
        return {}
    return {"movement_rollback_refs": _clear_movement_rollback_references(conn, ids)}


def _clear_movement_rollback_references(conn: Any, movement_ids: list[int]) -> int:
    if not movement_ids:
        return 0
    placeholders = ", ".join("?" for _ in movement_ids)
    row = conn.execute(
        f"SELECT COUNT(*) AS n FROM movements WHERE reverted_by_movement_id IN ({placeholders}) AND id NOT IN ({placeholders})",
        movement_ids + movement_ids,
    ).fetchone()
    conn.execute(
        f"UPDATE movements SET reverted_by_movement_id = NULL WHERE reverted_by_movement_id IN ({placeholders}) AND id NOT IN ({placeholders})",
        movement_ids + movement_ids,
    )
    return int(row["n"] or 0)


def _clean_delete_ids(raw_ids: Any) -> list[int]:
    if isinstance(raw_ids, str):
        raw_items = [item.strip() for item in raw_ids.replace("，", ",").split(",")]
    elif isinstance(raw_ids, list):
        raw_items = raw_ids
    else:
        raw_items = []
    ids: list[int] = []
    for raw in raw_items:
        if raw in ("", None):
            continue
        try:
            item_id = int(raw)
        except (TypeError, ValueError):
            raise ApiError(400, "记录 ID 必须是整数")
        if item_id <= 0:
            raise ApiError(400, "记录 ID 必须大于 0")
        if item_id not in ids:
            ids.append(item_id)
    if len(ids) > 100:
        raise ApiError(400, "单次最多删除 100 条记录")
    return ids


def _table_names(conn: Any) -> list[str]:
    rows = conn.execute(
        """
        SELECT name FROM sqlite_master
        WHERE type = 'table'
          AND name NOT LIKE 'sqlite_%'
        ORDER BY name
        """
    ).fetchall()
    return [str(row["name"]) for row in rows]


def _table_columns(conn: Any, table: str) -> list[str]:
    return [str(row["name"]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()]


def _required_columns(conn: Any, table: str) -> list[str]:
    columns = []
    for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall():
        if int(row["notnull"]) == 1 and row["dflt_value"] is None and int(row["pk"]) == 0:
            columns.append(str(row["name"]))
    return columns


def _primary_key_columns(conn: Any, table: str) -> list[str]:
    return [str(row["name"]) for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall() if int(row["pk"]) > 0]


def _quote_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _import_sheet(conn: Any, table: str, sheet: Any, mode: str) -> dict[str, Any]:
    allowed_columns = _table_columns(conn, table)
    if sheet.max_row < 1:
        raise ApiError(400, f"{sheet.title} 没有表头")
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    unknown = [header for header in headers if header and header not in allowed_columns]
    if unknown:
        raise ApiError(400, f"{table} 存在未知列：{', '.join(unknown)}")
    use_columns = [header for header in headers if header in allowed_columns]
    if not use_columns:
        raise ApiError(400, f"{table} 没有可导入列")
    auto_fill = {"created_at", "updated_at", "moved_at"}
    if table == "users":
        auto_fill.add("password_hash")
    missing_required = [col for col in _required_columns(conn, table) if col not in use_columns and col not in auto_fill]
    if missing_required:
        raise ApiError(400, f"{table} 缺少必填列：{', '.join(missing_required)}")
    pk_cols = _primary_key_columns(conn, table)
    if mode == "upsert" and not pk_cols:
        raise ApiError(400, f"{table} 没有主键，不能 upsert")
    if mode == "upsert" and not all(col in use_columns for col in pk_cols):
        raise ApiError(400, f"{table} upsert 需要主键列：{', '.join(pk_cols)}")

    result = {"success": 0, "inserted": 0, "updated": 0, "skipped": 0, "failed": []}
    timestamp = now_text()
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        record = {col: clean_excel_cell(row[idx], blank_as_none=True) for idx, col in enumerate(headers) if col in allowed_columns and idx < len(row)}
        record = {key: value for key, value in record.items() if value is not None}
        if not record:
            result["skipped"] += 1
            continue
        for col in ("created_at", "updated_at", "moved_at"):
            if col in allowed_columns and not record.get(col):
                record[col] = timestamp
        if table == "users" and "password_hash" in allowed_columns and not record.get("password_hash"):
            record["password_hash"] = hash_password(_initial_password())
        if mode == "append" and "id" in record and "id" in pk_cols:
            record.pop("id", None)
        try:
            if mode == "upsert":
                action = _upsert_record(conn, table, pk_cols, record)
                result[action] += 1
            else:
                _insert_record(conn, table, record)
                result["inserted"] += 1
            result["success"] += 1
        except Exception as exc:  # noqa: BLE001 - import reports per-row failures to the UI.
            result["failed"].append({"row": row_index, "error": str(exc), "record": record})
    return result


def _insert_record(conn: Any, table: str, record: dict[str, Any]) -> None:
    cols = list(record)
    placeholders = ", ".join("?" for _ in cols)
    conn.execute(f"INSERT INTO {_quote_ident(table)} ({', '.join(_quote_ident(col) for col in cols)}) VALUES ({placeholders})", [record[col] for col in cols])


def _upsert_record(conn: Any, table: str, pk_cols: list[str], record: dict[str, Any]) -> str:
    where = " AND ".join(f"{_quote_ident(col)} = ?" for col in pk_cols)
    exists = conn.execute(f"SELECT 1 FROM {_quote_ident(table)} WHERE {where} LIMIT 1", [record.get(col) for col in pk_cols]).fetchone()
    if exists:
        update_cols = [col for col in record if col not in pk_cols]
        if not update_cols:
            return "skipped"
        assignments = ", ".join(f"{_quote_ident(col)} = ?" for col in update_cols)
        params = [record[col] for col in update_cols] + [record.get(col) for col in pk_cols]
        conn.execute(f"UPDATE {_quote_ident(table)} SET {assignments} WHERE {where}", params)
        return "updated"
    else:
        _insert_record(conn, table, record)
        return "inserted"
